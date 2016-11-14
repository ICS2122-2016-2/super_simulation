import openpyxl
import pickle


def generar_histograma(escenario, path, benchmark=0, te=0, ti=0, m=0, optimizacion=0):

    if not benchmark and not optimizacion:
        with open('Output/{}/histograma_canasta_tentacion'.format(escenario), 'rb') as file:
            histograma_canasta_tentacion = pickle.load(file)
        with open('Output/{}/histograma_canasta_caminata'.format(escenario), 'rb') as file:
            histograma_canasta_caminata = pickle.load(file)
        with open('Output/{}/histograma_canasta_tiempo'.format(escenario), 'rb') as file:
            histograma_canasta_tiempo = pickle.load(file)
    if benchmark:
        histograma_canasta_tentacion = te
        histograma_canasta_tiempo = ti
        histograma_canasta_caminata = m
    if optimizacion:
        histograma_canasta_tentacion = te
        histograma_canasta_tiempo = ti
        histograma_canasta_caminata = m

    doc = openpyxl.Workbook()
    doc.create_sheet()
    doc.create_sheet()

    sheet = doc.worksheets[0]
    sheet.title = 'Canastas - Tiempo'
    cols = [1, 2]
    rows = list(range(1, len(histograma_canasta_tiempo)+1))
    for col in cols:
        for r in rows:
            sheet.cell(row=r, column=col, value=float(histograma_canasta_tiempo[r-1][col-1]))

    sheet = doc.worksheets[1]
    sheet.title = 'Canastas - Metros'
    cols = [1, 2]
    rows = list(range(1, len(histograma_canasta_caminata)+1))
    for col in cols:
        for r in rows:
            sheet.cell(row=r, column=col, value=float(histograma_canasta_caminata[r-1][col-1]))

    sheet = doc.worksheets[2]
    sheet.title = 'Canasta - Tentacion'
    cols = [1, 2]
    rows = list(range(1, len(histograma_canasta_tentacion)+1))

    for col in cols:
        for r in rows:
            sheet.cell(row=r, column=col, value=float(histograma_canasta_tentacion[r-1][col-1]))

    doc.save(path)